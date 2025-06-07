# excel_manager.py
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime # Using standard datetime for Gregorian dates 🗓️
import os
import jdatetime

def create_initial_excel(file_path):
    """
    Creates a new Excel file with 'Customers', 'Transactions', and 'Form' sheets
    and their respective headers if the file does not already exist. 🆕
    """
    if not os.path.exists(file_path):
        wb = Workbook()

        # Remove default sheet created by openpyxl 🧹
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Customers sheet with headers 👥
        ws_customers = wb.create_sheet("Customers")
        ws_customers.append(["کد مشتری", "نام", "شماره تماس", "تاریخ عضویت", "توضیحات"])

        # Create Transactions sheet with headers 💰
        ws_transactions = wb.create_sheet("Transactions")
        ws_transactions.append(["شناسه مشتری", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])

        # Create Form sheet with headers (primarily for mimicking the Excel input form structure) 📝
        ws_form = wb.create_sheet("Form")
        ws_form.append(["نام مشتری", "کد مشتری", "شماره تماس", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])

        wb.save(file_path)
        print(f"Initial Excel file created at {file_path} ✨")

def get_next_customer_id(df_customers):
    """
    Generates the next sequential customer ID (e.g., C001, C002). 🆔
    It finds the maximum existing numerical part of customer IDs and increments it.
    """
    if df_customers.empty:
        return "C001"
    else:
        # Extract numerical part, convert to int, find max, then format 🔢
        # Handle cases where 'کد مشتری' might not start with 'C' or might be non-numeric
        max_id = 0
        for customer_id in df_customers["کد مشتری"]:
            if isinstance(customer_id, str) and customer_id.startswith('C'):
                try:
                    num_part = int(customer_id[1:])
                    if num_part > max_id:
                        max_id = num_part
                except ValueError:
                    # Ignore invalid customer IDs 🚫
                    pass
        return f"C{max_id + 1:03d}" # Format to ensure leading zeros (e.g., C001) ➡️

def get_next_invoice_number(df_transactions):
    """
    Generates the next sequential invoice number (e.g., INV001, INV002). 🧾
    It finds the maximum existing numerical part of invoice numbers and increments it.
    """
    if df_transactions.empty:
        return "INV001"
    else:
        # Extract numerical part, convert to int, find max, then format 🔢
        max_inv_num = 0
        for inv_num in df_transactions["شماره فاکتور"]:
            if isinstance(inv_num, str) and inv_num.startswith('INV'):
                try:
                    num_part = int(inv_num[3:])
                    if num_part > max_inv_num:
                        max_inv_num = num_part
                except ValueError:
                    # Ignore invalid invoice numbers 🚫
                    pass
        return f"INV{max_inv_num + 1:03d}" # Format to ensure leading zeros (e.g., INV001) ➡️

def save_purchase(file_path, customer_name, customer_phone, amount):
    """
    Saves a new purchase record to the Excel file. 💾
    This involves updating the 'Customers' and 'Transactions' sheets,
    and optionally the 'Form' sheet for display purposes.
    """
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: Excel file not found at {file_path}. Creating a new one. ⚠️")
        create_initial_excel(file_path) # Create if not found ➕
        wb = load_workbook(file_path) # Load again 🔄

    # --- Update Customers Sheet 👥 ---
    ws_customers = wb["Customers"]
    # Read existing customer data into a DataFrame for easier lookup 🔍
    df_customers = pd.DataFrame(ws_customers.iter_rows(min_row=2, values_only=True), 
                                columns=["کد مشتری", "نام", "شماره تماس", "تاریخ عضویت", "توضیحات"])

    customer_id = None
    # Check if customer already exists by phone number ✅
    existing_customer = df_customers[df_customers["شماره تماس"] == customer_phone]

    if not existing_customer.empty:
        # Customer exists, use their existing ID 👍
        customer_id = existing_customer["کد مشتری"].iloc[0]
        print(f"Existing customer found: {customer_id} 🧑‍🤝‍🧑")
        # Optional: Update customer name if it has changed (good practice) 🔄
        for row_idx, row in enumerate(ws_customers.iter_rows(min_row=2), start=2):
            if row[2].value == customer_phone: # Column C (index 2) is "شماره تماس"
                if row[1].value != customer_name: # Column B (index 1) is "نام"
                    ws_customers.cell(row=row_idx, column=2, value=customer_name)
                    print(f"Updated name for customer {customer_id} to {customer_name} ✅")
                break
    else:
        # New customer, generate a new ID and add to Customers sheet 🆕
        customer_id = get_next_customer_id(df_customers)
        # Get current date in Gregorian format YYYY/MM/DD 🗓️
        today_date = jdatetime.date.today().strftime("%Y-%m-%d")
        ws_customers.append([customer_id, customer_name, customer_phone, today_date, ""]) # Add empty description ➕
        print(f"Added new customer: {customer_id} ({customer_name}, {customer_phone}) 🎉")

    # --- Update Transactions Sheet 💰 ---
    ws_transactions = wb["Transactions"]
    # Read existing transaction data to generate the next invoice number 🔍
    df_transactions = pd.DataFrame(ws_transactions.iter_rows(min_row=2, values_only=True), 
                                   columns=["شناسه مشتری", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])

    invoice_number = get_next_invoice_number(df_transactions)
    # Current date in Gregorian for invoice 🗓️
    invoice_date = jdatetime.date.today().strftime("%Y-%m-%d")

    ws_transactions.append([customer_id, invoice_date, invoice_number, amount])
    print(f"Added new transaction: {invoice_number} for customer {customer_id} with amount {amount} 💸")

    # --- Update Form Sheet (for display/mimicking original Excel behavior) 📝 ---
    # In a bot, this sheet isn't used for input, but we can update it with the last recorded transaction
    # to show the user what was just processed, if they were to open the Excel file. 📊
    ws_form = wb["Form"]
    # Clear previous data in Form sheet (assuming it only holds one transaction at a time) 🧹
    for row in ws_form.iter_rows(min_row=2, max_row=ws_form.max_row, max_col=ws_form.max_column):
        for cell in row:
            cell.value = None
    # Append the new transaction details to the Form sheet ➕
    ws_form.append([customer_name, customer_id, customer_phone, invoice_date, invoice_number, amount])
    print("Form sheet updated with latest transaction. ✅")

    # Save the entire workbook 💾
    wb.save(file_path)
    print(f"Excel file saved successfully at {file_path} ✨")


def save_purchase_bulk(file_path, customer_name, customer_phone, amount, description=""):
    """
    Saves bulk customer data into the Excel file. Handles new and existing customers.
    """
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        create_initial_excel(file_path)
        wb = load_workbook(file_path)

    # --- Customers Sheet ---
    ws_customers = wb["Customers"]
    df_customers = pd.DataFrame(ws_customers.iter_rows(min_row=2, values_only=True),
                                columns=["کد مشتری", "نام", "شماره تماس", "تاریخ عضویت", "توضیحات"])

    # Check if customer already exists
    existing = df_customers[df_customers["شماره تماس"] == customer_phone]
    if not existing.empty:
        customer_id = existing["کد مشتری"].iloc[0]
    else:
        customer_id = get_next_customer_id(df_customers)
        today_date = jdatetime.date.today().strftime("%Y-%m-%d")
        ws_customers.append([customer_id, customer_name, customer_phone, today_date, description])

    # --- Transactions Sheet ---
    ws_transactions = wb["Transactions"]
    df_transactions = pd.DataFrame(ws_transactions.iter_rows(min_row=2, values_only=True),
                                   columns=["شناسه مشتری", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])

    invoice_number = get_next_invoice_number(df_transactions)
    invoice_date = jdatetime.date.today().strftime("%Y-%m-%d")
    ws_transactions.append([customer_id, invoice_date, invoice_number, amount])

    # Form Sheet: optional
    ws_form = wb["Form"]
    for row in ws_form.iter_rows(min_row=2, max_row=ws_form.max_row):
        for cell in row:
            cell.value = None
    ws_form.append([customer_name, customer_id, customer_phone, invoice_date, invoice_number, amount])

    wb.save(file_path)


def get_customers_data(file_path):
    """
    Reads the 'Customers' sheet from the Excel file and returns it as a pandas DataFrame. 👥
    Additionally, calculates 'Total Transactions' and 'Total Spend' from the 'Transactions' sheet
    and merges them with the customer data.
    Returns an empty DataFrame if the file or sheet is not found, or an error occurs. 🚫
    """
    try:
        # Read Customers sheet 🏷️
        df_customers = pd.read_excel(file_path, sheet_name="Customers", header=0)
    except FileNotFoundError:
        print(f"Error: Excel file not found at {file_path}. ⚠️")
        return pd.DataFrame(columns=["کد مشتری", "نام", "شماره تماس", "تاریخ عضویت", "توضیحات"])
    except Exception as e:
        print(f"Error reading Customers sheet from {file_path}: {e} ❌")
        return pd.DataFrame(columns=["کد مشتری", "نام", "شماره تماس", "تاریخ عضویت", "توضیحات"])

    try:
        # Read Transactions sheet 💰
        df_transactions = pd.read_excel(file_path, sheet_name="Transactions", header=0)
    except Exception as e:
        print(f"Warning: Could not read Transactions sheet from {file_path}. Assuming no transactions. {e} ⚠️")
        df_transactions = pd.DataFrame(columns=["شناسه مشتری", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])
    
    # Calculate Total Transactions and Total Spend from transactions data 📈
    if not df_transactions.empty:
        # Ensure 'مبلغ (تومان)' is numeric for sum calculation
        df_transactions['مبلغ (تومان)'] = pd.to_numeric(df_transactions['مبلغ (تومان)'], errors='coerce').fillna(0)

        customer_summary = df_transactions.groupby('شناسه مشتری').agg(
            Total_Transactions=('شماره فاکتور', 'count'), # Count of transactions
            Total_Spend=('مبلغ (تومان)', 'sum') # Sum of amounts
        ).reset_index()
        
        # Merge this summary with the customers DataFrame
        # 'شناسه مشتری' from transactions matches 'کد مشتری' in customers
        df_customers = pd.merge(
            df_customers,
            customer_summary,
            left_on='کد مشتری',
            right_on='شناسه مشتری',
            how='left'
        )
        # Fill NaN values for customers with no transactions
        df_customers['Total_Transactions'].fillna(0, inplace=True)
        df_customers['Total_Spend'].fillna(0, inplace=True)

        # Drop the redundant 'شناسه مشتری' column from the merge
        if 'شناسه مشتری' in df_customers.columns:
            df_customers.drop(columns=['شناسه مشتری'], inplace=True)
        
        # Rename the new columns to Persian names as requested
        df_customers.rename(columns={
            'Total_Transactions': 'تعداد کل تراکنش‌ها',
            'Total_Spend': 'مجموع مبلغ خریدها'
        }, inplace=True)

    else:
        # If no transactions data, add these columns with default values (0)
        df_customers['تعداد کل تراکنش‌ها'] = 0
        df_customers['مجموع مبلغ خریدها'] = 0


    return df_customers

def get_transactions_data(file_path):
    """
    Reads the 'Transactions' sheet from the Excel file and returns it as a pandas DataFrame. 💰
    Returns an empty DataFrame if the file or sheet is not found, or an error occurs. 🚫
    """
    try:
        # Use header=0 to indicate the first row is the header 🏷️
        df = pd.read_excel(file_path, sheet_name="Transactions", header=0)
        return df
    except FileNotFoundError:
        print(f"Error: Excel file not found at {file_path}. ⚠️")
        return pd.DataFrame(columns=["شناسه مشتری", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])
    except Exception as e:
        print(f"Error reading Transactions sheet from {file_path}: {e} ❌")
        return pd.DataFrame(columns=["شناسه مشتری", "تاریخ فاکتور", "شماره فاکتور", "مبلغ (تومان)"])

def create_temp_excel_report(df: pd.DataFrame, sheet_name: str, report_type: str, data_dir: str) -> str:
    """
    Creates a temporary Excel file containing the given DataFrame in a single sheet. 📊
    The filename includes the user ID, report type, and current date (Gregorian). 🗓️

    Args:
        df (pd.DataFrame): The DataFrame to save to the Excel file. 📈
        sheet_name (str): The name of the sheet in the new Excel file. 🏷️
        user_id (int): The Telegram user ID, used for naming the file. 🆔
        report_type (str): Type of report (e.g., "customers", "transactions") for filename. 📝
        data_dir (str): The directory where temporary files should be stored. 📁

    Returns:
        str: The full path to the created temporary Excel file. 📂
    """
    # Use standard datetime for the date in the temporary file name 🗓️
    today_date_str = jdatetime.date.today().strftime("%Y-%m-%d")
    # Example filename: user_data/12345_customers_2024-06-03.xlsx 📄
    temp_file_name = f"{report_type}_{today_date_str}.xlsx"
    temp_file_path = os.path.join(data_dir, temp_file_name)

    wb = Workbook()
    # Remove default sheet 🧹
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws = wb.create_sheet(sheet_name)

    # Write headers ✍️
    ws.append(df.columns.tolist())
    # Write data rows ➕
    for r_idx, row in df.iterrows():
        ws.append(row.tolist())

    wb.save(temp_file_path)
    print(f"Temporary Excel report created at: {temp_file_path} ✨")
    return temp_file_path

